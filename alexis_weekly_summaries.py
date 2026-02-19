"""
alexis_weekly_summaries.py

Reads all ALEXIS weekly snapshots ({YYYY}_{mon}_w{N}.json) from a folder
and produces a single Excel workbook with one sheet per summary:

    Trial Counts            ← drug_trial_counts
    Drug Overview           ← drug_info_overview
    Modality Provenance     ← drug_modality_provenance
    TA Provenance           ← drug_ta_provenance
    Trials by TA            ← ta_modality_counts_true_drugs (summed per TA)
    Non-Disease Categories  ← non_disease_study_categories (dynamic keys, unioned)
    Mesh Missing            ← drug_mesh_missing_condition

Weeks as rows (newest first).  Sort order comes from the filename, not
file contents.  The Week column displays metadata.as_of.

Usage:
    python alexis_weekly_summaries.py                        # current dir, output here
    python alexis_weekly_summaries.py /path/to/snapshots     # specific input folder
    python alexis_weekly_summaries.py /in /out/report.xlsx   # specific input + output
"""

import json, re, glob, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config ────────────────────────────────────────────────────────────
MONTH_MAP = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
    'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
}

HEADER_FILL  = PatternFill('solid', fgColor='1F3864')
HEADER_FONT  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
DATE_FONT    = Font(name='Arial', size=10, bold=True)
DATA_FONT    = Font(name='Arial', size=10)
ALT_ROW_FILL = PatternFill('solid', fgColor='D6E4F0')
THIN_BORDER  = Border(
    left=Side('thin', color='B4C6E5'),
    right=Side('thin', color='B4C6E5'),
    top=Side('thin', color='B4C6E5'),
    bottom=Side('thin', color='B4C6E5'),
)
NUM_FMT = '#,##0'

# ── helpers ───────────────────────────────────────────────────────────
def parse_filename(fname):
    m = re.match(r'^(?:reclassified_)?(\d{4})_([a-z]{3})_w(\d+)\.json$', os.path.basename(fname))
    if not m:
        return None
    year, mon, week = int(m.group(1)), m.group(2), int(m.group(3))
    month_num = MONTH_MAP.get(mon)
    return (year, month_num, week) if month_num else None

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = THIN_BORDER

def style_data_row(ws, row, max_col, is_alt):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        if col == 1:
            cell.font      = DATE_FONT
            cell.alignment = Alignment(horizontal='left')
        elif col == 2:
            cell.font      = Font(name='Arial', size=10, color='808080')
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.font          = DATA_FONT
            cell.number_format = NUM_FMT
            cell.alignment     = Alignment(horizontal='right')
        if is_alt:
            cell.fill = ALT_ROW_FILL

# ── first pass: load all snapshots ────────────────────────────────────
def load_snapshots(folder):
    parsed_files = []
    for f in glob.glob(os.path.join(folder, '*.json')):
        sk = parse_filename(f)
        if sk:
            parsed_files.append((sk, f))
    parsed_files.sort(key=lambda x: x[0])

    rows = []
    all_nd_categories = set()

    for sort_key, fname in parsed_files:
        with open(fname) as fh:
            data = json.load(fh)
        meta    = data['metadata']
        summary = data['summary']

        ta_totals = {
            ta: sum(mod_dict.values())
            for ta, mod_dict in summary.get('ta_modality_counts_true_drugs', {}).items()
        }
        nd_cats = summary.get('non_disease_study_categories', {})
        all_nd_categories.update(nd_cats.keys())

        rows.append({
            'sort_key':                     sort_key,
            'as_of':                        meta.get('as_of'),
            'run_id':                       meta.get('run_id'),
            'drug_trial_counts':            summary.get('drug_trial_counts', {}),
            'drug_info_overview':           summary.get('drug_info_overview', {}),
            'drug_modality_provenance':     summary.get('drug_modality_provenance', {}),
            'drug_ta_provenance':           summary.get('drug_ta_provenance', {}),
            'ta_totals':                    ta_totals,
            'non_disease_study_categories': nd_cats,
            'mesh_missing':                 summary.get('drug_mesh_missing_condition', {}).get('mesh_missing_condition', 0),
        })

    rows.sort(key=lambda r: r['sort_key'], reverse=True)   # newest first

    all_tas       = sorted({ta  for r in rows for ta  in r['ta_totals']})
    all_nd_sorted = sorted(all_nd_categories)

    return rows, all_tas, all_nd_sorted

# ── second pass: write workbook ───────────────────────────────────────
def build_workbook(rows, all_tas, all_nd_sorted):
    wb = Workbook()

    def write_flat_sheet(sheet_name, extract_fn, col_headers):
        ws      = wb.create_sheet(sheet_name)
        headers = ['Week', 'Run ID'] + col_headers
        max_col = len(headers)

        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header_row(ws, 1, max_col)
        ws.row_dimensions[1].height = 22

        for i, r in enumerate(rows):
            excel_row = i + 2
            ws.cell(row=excel_row, column=1, value=r['as_of'])
            ws.cell(row=excel_row, column=2, value=r['run_id'])
            for c, val in enumerate(extract_fn(r), 3):
                ws.cell(row=excel_row, column=c, value=val if val else None)
            style_data_row(ws, excel_row, max_col, is_alt=(i % 2 == 1))

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        for idx in range(3, max_col + 1):
            ws.column_dimensions[get_column_letter(idx)].width = max(len(col_headers[idx-3]) + 2, 16)

    # Trial Counts
    KEYS_TC = ['total_trials','drug_trials','non_drug_trials','drug_trials_with_unknown_modality']
    HDRS_TC = ['Total Trials','Drug Trials','Non-Drug Trials','Unknown Modality']
    write_flat_sheet('Trial Counts', lambda r: [r['drug_trial_counts'].get(k,0) for k in KEYS_TC], HDRS_TC)

    # Drug Overview
    KEYS_DO = ['drug_trials_total','drug_trials_with_info_flags','drug_trials_with_unknown_modality','drug_trials_unknown_with_info']
    HDRS_DO = ['Drug Trials Total','With Info Flags','Unknown Modality','Unknown + Info']
    write_flat_sheet('Drug Overview', lambda r: [r['drug_info_overview'].get(k,0) for k in KEYS_DO], HDRS_DO)

    # Modality Provenance
    write_flat_sheet('Modality Provenance',
        lambda r: [r['drug_modality_provenance'].get(k,0) for k in ['mesh','text_fallback','base_only']],
        ['MeSH','Text Fallback','Base Only'])

    # TA Provenance
    write_flat_sheet('TA Provenance',
        lambda r: [r['drug_ta_provenance'].get(k,0) for k in ['mesh','text_fallback','multi_ta_mesh']],
        ['MeSH','Text Fallback','Multi-TA MeSH'])

    # Trials by TA
    write_flat_sheet('Trials by TA',
        lambda r: [r['ta_totals'].get(ta,0) for ta in all_tas],
        all_tas)

    # Non-Disease Categories (dynamic keys — union across all weeks)
    write_flat_sheet('Non-Disease Categories',
        lambda r: [r['non_disease_study_categories'].get(cat,0) for cat in all_nd_sorted],
        all_nd_sorted)

    # Mesh Missing
    write_flat_sheet('Mesh Missing',
        lambda r: [r['mesh_missing']],
        ['Mesh Missing Count'])

    del wb[wb.sheetnames[0]]   # remove default sheet
    return wb


# ── main ──────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) >= 3:
        folder  = sys.argv[1]
        out     = sys.argv[2]
    elif len(sys.argv) == 2:
        folder  = sys.argv[1]
        out     = os.path.join(folder, 'alexis_weekly_summaries.xlsx')
    else:
        folder  = '.'
        out     = 'alexis_weekly_summaries.xlsx'

    rows, all_tas, all_nd_sorted = load_snapshots(folder)

    if not rows:
        print(f"No valid snapshots found in {folder}")
        sys.exit(1)

    print(f"Loaded {len(rows)} snapshots from {folder}")
    print(f"  TAs:                {len(all_tas)}")
    print(f"  Non-disease cats:   {len(all_nd_sorted)}")

    wb = build_workbook(rows, all_tas, all_nd_sorted)
    wb.save(out)

    print(f"Saved {out}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")